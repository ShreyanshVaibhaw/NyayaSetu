"""Generate interest calculation Excel sheets."""

from __future__ import annotations

from io import BytesIO
from typing import List

import pandas as pd
from openpyxl.utils import get_column_letter

from src.common.models import InterestCalculation


def build_interest_calculation_excel(calculation: InterestCalculation) -> bytes:
    """Create XLSX workbook with summary formulas and month-wise breakdown."""
    rows: List[dict] = calculation.calculation_breakdown
    df = pd.DataFrame(rows if rows else [{"note": "No monthly breakdown available for this date range."}])
    buf = BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Breakdown", index=False)
        wb = writer.book
        ws = wb.create_sheet("Summary")

        ws["A1"] = "Metric"
        ws["B1"] = "Value"
        ws["A2"] = "Invoice"
        ws["B2"] = calculation.invoice_number
        ws["A3"] = "Principal"
        ws["B3"] = calculation.principal
        ws["A4"] = "Interest (Formula)"
        ws["A5"] = "Total Due (Formula)"
        ws["A6"] = "Applicable Annual Rate (%)"
        ws["B6"] = calculation.applicable_rate

        if rows:
            interest_col = get_column_letter(df.columns.get_loc("interest_component") + 1)
            closing_col = get_column_letter(df.columns.get_loc("closing_balance") + 1)
            end_row = len(df) + 1
            ws["B4"] = f"=SUM(Breakdown!{interest_col}2:{interest_col}{end_row})"
            ws["B5"] = f"=Breakdown!{closing_col}{end_row}"
        else:
            ws["B4"] = calculation.interest_amount
            ws["B5"] = calculation.total_due

        for cell in ("B3", "B4", "B5"):
            ws[cell].number_format = "#,##0.00"

    return buf.getvalue()
